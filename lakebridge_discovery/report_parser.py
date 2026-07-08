"""
Defensive mapper: Lakebridge Analyzer report (JSON preferred, Excel
fallback) -> lakebridge_discovery.schema.LakebridgeDiscoveryResult.

For JSON reports, the *object inventory and dependency* shape is in fact
publicly known: Lakebridge's `analyze` command delegates to the
`databricks-labs-bladespector` package, whose installed JSON schemas
(databricks/labs/bladespector/schemas/analyzer-{sql,etl}-schema.json) define
`inventory[].objectRel[]` (per-object {object, action, count} relationships),
ETL's `inventory[].sqlStatements[].objectRel[]`, and ETL's top-level
`subJobInfo[]` (parent/child job relationships) as the native dependency
data -- see extract_native_report_dependencies() below, which is preferred
over dependency_extractor.py's regex text-scan (that module only fills gaps
for objects this function found zero edges for).

Everything else here (_CATEGORY_KEYWORDS, _apply_rows, and Excel handling)
remains keyword-based best-effort classification, since the Excel report's
sheet layout and any *other* JSON shapes (older/newer Analyzer versions)
aren't pinned down the same way -- never raises on an unrecognized shape,
logs a warning and moves on. Re-validate against a real report the first
time this runs against an actual Databricks workspace, and tighten further.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from lakebridge_discovery.dependency_extractor import _build_inventory, _normalize_ref, _resolve
from lakebridge_discovery.logging_setup import logger
from lakebridge_discovery.schema import (
    AnalyzeInvocationEntity,
    LakebridgeDependencyRef,
    LakebridgeDiscoveryResult,
    LakebridgeObjectRef,
)

# Bladespector's objectRel "action" enum -> this engine's relationship_type
# vocabulary. "create"/"drop" are DDL on the object itself (the script
# defining/dropping its own object), not a cross-object edge -- intentionally
# unmapped, so those actions are skipped rather than emitting a self-loop.
_ACTION_TO_RELATIONSHIP = {
    "read": "reads", "source": "reads", "lookup": "reads",
    "write": "writes", "target": "writes",
    "execute": "calls",
}

# Ordered so more specific keywords (e.g. "stored procedure") are checked
# before generic ones. Matched against lower-cased sheet names / JSON key
# paths / row "type"-ish field values.
_CATEGORY_KEYWORDS: list[tuple[str, str]] = [
    ("dependenc", "dependency"),
    ("unsupported", "unsupported"),
    ("stored procedure", "stored_procedure"),
    ("procedure", "stored_procedure"),
    ("proc", "stored_procedure"),
    ("view", "view"),
    ("function", "function"),
    ("trigger", "trigger"),
    ("synonym", "synonym"),
    ("schema", "schema"),
    ("package", "package"),
    ("workflow", "package"),
    ("job", "package"),
    ("table", "table"),
]

_NAME_KEYS = [
    "name", "object_name", "table_name", "view_name", "procedure_name",
    "function_name", "trigger_name", "package_name", "job_name", "object",
]
_COMPLEXITY_KEYS = ["complexity", "complexity_score", "effort", "complexity_level"]
_SOURCE_KEYS = ["source_object", "source", "from", "from_object", "parent"]
_TARGET_KEYS = ["target_object", "target", "to", "to_object", "child"]
_RELATIONSHIP_KEYS = ["relationship_type", "relationship", "type", "dependency_type"]


def _classify(label: str) -> str | None:
    lowered = label.lower()
    for keyword, category in _CATEGORY_KEYWORDS:
        if keyword in lowered:
            return category
    return None


def _first_present(row: dict, keys: list[str]) -> str | None:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return str(row[key])
    return None


def _row_to_object_ref(row: dict, category: str, source_tech: str, raw_category: str) -> LakebridgeObjectRef:
    name = _first_present(row, _NAME_KEYS) or json.dumps(row, default=str)[:80]
    return LakebridgeObjectRef(
        object_type=category,
        name=name,
        source_tech=source_tech,
        raw_category=raw_category,
        complexity=_first_present(row, _COMPLEXITY_KEYS),
    )


def _row_to_dependency_ref(row: dict, raw_category: str) -> LakebridgeDependencyRef:
    # Analyzer-report-derived rows carry no reliable object-type info, unlike
    # dependency_extractor.py's own text-scan edges -- tagged distinctly and
    # left unresolved rather than defaulting to a misleading resolved=True.
    return LakebridgeDependencyRef(
        source_object=_first_present(row, _SOURCE_KEYS) or "(unknown)",
        target_object=_first_present(row, _TARGET_KEYS) or "(unknown)",
        relationship_type=_first_present(row, _RELATIONSHIP_KEYS) or "unknown",
        raw_category=raw_category,
        discovery_method="lakebridge_report",
        resolved=False,
    )


def _apply_rows(result: LakebridgeDiscoveryResult, rows: list[dict], raw_category: str, source_tech: str) -> int:
    category = _classify(raw_category)
    if category is None:
        # Try classifying by an in-row "type"/"category" field instead.
        applied = 0
        for row in rows:
            if not isinstance(row, dict):
                continue
            type_hint = _first_present(row, ["type", "object_type", "category"]) or ""
            row_category = _classify(type_hint)
            if row_category is None:
                continue
            applied += _apply_rows(result, [row], row_category, source_tech)
        return applied

    count = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        if category == "dependency":
            result.dependencies.append(_row_to_dependency_ref(row, raw_category))
        elif category == "unsupported":
            result.unsupported_objects.append(_row_to_object_ref(row, "unsupported", source_tech, raw_category))
        else:
            bucket: list[LakebridgeObjectRef] = getattr(result, _PLURAL.get(category, category), None)
            if bucket is None:
                continue
            bucket.append(_row_to_object_ref(row, category, source_tech, raw_category))
        count += 1
    return count


_PLURAL = {
    "table": "tables", "view": "views", "stored_procedure": "stored_procedures",
    "function": "functions", "trigger": "triggers", "synonym": "synonyms",
    "schema": "schemas", "package": "packages",
}

# Verified against a real MS SQL Server Analyzer report (run 2026-07-03): the
# object inventory isn't in anything the generic keyword classifier below can
# find -- it's keyed by program/file name (Excel: "SQL Programs"."Program
# Name", JSON: top-level "inventory"[]."name"), which already carries the
# object type via source_exporter.py's own `{kind}__{schema}.{name}.ext` file
# naming convention (e.g. "table__Sales.Store.sql"). Other sheets/keys are
# usage/complexity aggregates, not object catalogs -- classifying them by
# name (e.g. Excel sheet "Functions") misreads built-in-function call counts
# as a list of user-defined functions.
_PROGRAM_NAME_HEADER = "Program Name"
_NON_INVENTORY_SHEETS = {
    "Functions", "Functions by Script", "Scripts Functions Xref",
    "SQL Script Categories", "UNKNOWN SQL Category", "SQL Special Patterns",
    "SQL Data Types", "Summary",
}


def _parse_program_name(program_name: str) -> tuple[str, str] | None:
    """Decomposes source_exporter.py's `{kind}__{schema}.{name}.ext` file
    naming convention into (object_type, "schema.name"). Returns None if the
    name doesn't follow that convention, or classifies as dependency/
    unsupported (not a real object category)."""
    prefix, sep, rest = str(program_name).partition("__")
    if not sep:
        return None
    category = _classify(prefix.replace("_", " "))
    if category is None or category in ("dependency", "unsupported"):
        return None
    name = rest.rsplit(".", 1)[0] if "." in rest else rest
    return category, name


# SSIS inventory rows (unlike the SQL-side rows _parse_program_name handles)
# don't follow source_exporter.py's "{kind}__{name}" file-naming convention --
# each row already carries its own "type" field instead (confirmed against a
# real SSIS Analyzer report: every entry has "type": "Package" and a plain
# "name" like "Pkg_ArchiveOldData", no "__" separator at all), so
# _parse_program_name always returns None for them. This is the fallback
# classification path for exactly that case.
_ROW_TYPE_TO_CATEGORY = {
    "Package": "package",
}


def _classify_by_row_type(row: dict) -> str | None:
    row_type = row.get("type")
    return _ROW_TYPE_TO_CATEGORY.get(row_type) if isinstance(row_type, str) else None


def _apply_program_inventory_rows(
    result: LakebridgeDiscoveryResult, rows: list[dict], source_tech: str, name_field: str, complexity_field: str
) -> int:
    count = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        program_name = row.get(name_field)
        if not program_name:
            continue
        parsed = _parse_program_name(program_name)
        if parsed is not None:
            category, name = parsed
        else:
            category = _classify_by_row_type(row)
            if category is None:
                continue
            name = str(program_name)
        bucket = getattr(result, _PLURAL.get(category, category), None)
        if bucket is None:
            continue
        bucket.append(LakebridgeObjectRef(
            object_type=category,
            name=name,
            source_tech=source_tech,
            raw_category=name_field,
            complexity=row.get(complexity_field) or _first_present(row, _COMPLEXITY_KEYS),
        ))
        count += 1
    return count


# Names that show up as an "object"/"target"/lineage source but aren't real
# catalog objects -- SQL Server's trigger-only virtual tables, and any
# "@variable" (e.g. a table-valued function's own return-value variable, see
# extract_native_report_dependencies' object_lineage handling below). An
# edge naming one of these as an endpoint conveys nothing migration-relevant
# (nearly every trigger "reads inserted"; a function trivially "writes" its
# own return variable) so these are skipped rather than emitted as unresolved.
_PSEUDO_OBJECT_NAMES = {"inserted", "deleted"}


def _is_pseudo_object(name: str) -> bool:
    return name.startswith("@") or name.strip().lower() in _PSEUDO_OBJECT_NAMES


def _emit_dependency(
    result: LakebridgeDiscoveryResult, qualified: dict[str, str], bare_index: dict[str, str], seen_edges: set[tuple],
    source_object: str, source_type: str, own_schema: str | None,
    target_raw: str | None, relationship_type: str | None, raw_category: str,
) -> None:
    if relationship_type is None or not target_raw or _is_pseudo_object(str(target_raw)):
        return
    bare_name, schema = _normalize_ref(str(target_raw))
    if not bare_name:
        return
    matched_key, target_type = _resolve(bare_name, schema, own_schema, qualified, bare_index)
    target_object = matched_key or (f"{schema}.{bare_name}" if schema else bare_name)
    if target_object.lower() == source_object.lower():
        return  # self-loop (e.g. a CREATE statement's own object) -- not a real dependency
    edge_key = (source_object, target_object, relationship_type)
    if edge_key in seen_edges:
        return
    seen_edges.add(edge_key)
    result.dependencies.append(LakebridgeDependencyRef(
        source_object=source_object, target_object=target_object, relationship_type=relationship_type,
        raw_category=raw_category, source_type=source_type, target_type=target_type,
        discovery_method="lakebridge_report", resolved=matched_key is not None,
    ))


def _extract_object_rel_dependencies(
    data: dict, result: LakebridgeDiscoveryResult, qualified: dict[str, str], bare_index: dict[str, str], seen_edges: set[tuple],
) -> None:
    """Newer Bladespector shape: per-inventory-item `objectRel[]` /
    `sqlStatements[].objectRel[]`, plus top-level `subJobInfo[]` (ETL job
    hierarchy). Confirmed absent from at least one real installed Analyzer
    version (which uses object_lineage instead, see
    _extract_object_lineage_dependencies) -- kept so either shape works."""
    inventory = data.get("inventory")
    if not isinstance(inventory, list):
        return

    for item in inventory:
        if not isinstance(item, dict):
            continue
        parsed = _parse_program_name(item.get("name", ""))
        if parsed is None:
            continue
        source_type, source_object = parsed
        own_schema = source_object.split(".", 1)[0].lower() if "." in source_object else None

        rel_lists = [item.get("objectRel") or []]
        for stmt in item.get("sqlStatements") or []:
            if isinstance(stmt, dict):
                rel_lists.append(stmt.get("objectRel") or [])

        for raw_category, rel_list in zip(("objectRel", "sqlStatements.objectRel"), rel_lists):
            for rel in rel_list:
                if not isinstance(rel, dict):
                    continue
                relationship_type = _ACTION_TO_RELATIONSHIP.get(rel.get("action"))
                _emit_dependency(
                    result, qualified, bare_index, seen_edges, source_object, source_type, own_schema,
                    rel.get("object"), relationship_type, raw_category,
                )

    for edge in data.get("subJobInfo") or []:
        if not isinstance(edge, dict):
            continue
        parent, child = edge.get("parent"), edge.get("child")
        if not parent or not child or parent == child:
            continue
        edge_key = (parent, child, "calls")
        if edge_key in seen_edges:
            continue
        seen_edges.add(edge_key)
        result.dependencies.append(LakebridgeDependencyRef(
            source_object=parent, target_object=child, relationship_type="calls",
            raw_category="subJobInfo", source_type="package", target_type="package",
            discovery_method="lakebridge_report", resolved=True,
        ))


def _extract_object_lineage_dependencies(
    data: dict, result: LakebridgeDiscoveryResult, qualified: dict[str, str], bare_index: dict[str, str], seen_edges: set[tuple],
) -> None:
    """Older/other Bladespector shape (confirmed present in a real installed
    Analyzer's JSON, 2026-07-06 run against AdventureWorks2022): object-
    centric `object_lineage[]`, each entry `{target, reads?: [file...],
    writes?: [{file, sources: [object...]}]}`.

    `reads` lists files that read FROM target -> file --reads--> target.
    Each `writes` entry is a file that writes INTO target, itself sourced
    from `sources` -> file --writes--> target, and file --reads--> each
    source (this is the richer half: it gives per-write provenance, e.g. a
    trigger's write to one table sourced from a *different* table it joins
    in, which a plain regex FROM/JOIN scan of the trigger body would also
    find, but object_lineage already gives it to us pre-resolved).
    `target`/`sources` entries that are pseudo-objects (SQL Server's
    inserted/deleted trigger tables, a function's own "@returnVar") are
    filtered by _emit_dependency, not here.
    """
    for entry in data.get("object_lineage") or []:
        if not isinstance(entry, dict):
            continue
        target = entry.get("target")
        if not target:
            continue

        for file_name in entry.get("reads") or []:
            parsed = _parse_program_name(file_name) if file_name else None
            if parsed is None:
                continue
            source_type, source_object = parsed
            own_schema = source_object.split(".", 1)[0].lower() if "." in source_object else None
            _emit_dependency(
                result, qualified, bare_index, seen_edges, source_object, source_type, own_schema,
                target, "reads", "object_lineage",
            )

        for write in entry.get("writes") or []:
            if not isinstance(write, dict):
                continue
            parsed = _parse_program_name(write.get("file", "")) if write.get("file") else None
            if parsed is None:
                continue
            source_type, source_object = parsed
            own_schema = source_object.split(".", 1)[0].lower() if "." in source_object else None
            _emit_dependency(
                result, qualified, bare_index, seen_edges, source_object, source_type, own_schema,
                target, "writes", "object_lineage",
            )
            for source_name in write.get("sources") or []:
                _emit_dependency(
                    result, qualified, bare_index, seen_edges, source_object, source_type, own_schema,
                    source_name, "reads", "object_lineage",
                )


def extract_native_report_dependencies(data: dict, result: LakebridgeDiscoveryResult, source_tech: str) -> None:
    """Extracts dependency edges directly from the Analyzer's own JSON --
    the Analyzer already computes per-object relationships, so this is the
    primary dependency source; dependency_extractor.py's regex scan only
    fills gaps for objects this leaves with zero edges.

    Analyzer-version agnostic: tries both known native shapes
    unconditionally (a given report will realistically only populate one,
    but nothing breaks if both happened to be present -- overlapping edges
    just dedupe against the same `seen_edges` set):
      - object_lineage[] (see _extract_object_lineage_dependencies)
      - objectRel[] / sqlStatements[].objectRel[] / subJobInfo[] (see
        _extract_object_rel_dependencies)
    """
    qualified, bare_index = _build_inventory(result)
    seen_edges: set[tuple] = set()
    _extract_object_lineage_dependencies(data, result, qualified, bare_index, seen_edges)
    _extract_object_rel_dependencies(data, result, qualified, bare_index, seen_edges)


def _walk_json_for_lists(node: Any, path: str, result: LakebridgeDiscoveryResult, source_tech: str) -> int:
    """Recursively finds every list-of-dicts in the JSON report and applies
    it under whatever key name led to it -- handles reports shaped as a flat
    dict of category->rows, or nested under a wrapper object."""
    applied = 0
    if isinstance(node, dict):
        for key, value in node.items():
            applied += _walk_json_for_lists(value, key, result, source_tech)
    elif isinstance(node, list):
        if node and isinstance(node[0], dict):
            applied += _apply_rows(result, node, path, source_tech)
        else:
            for item in node:
                applied += _walk_json_for_lists(item, path, result, source_tech)
    return applied


#  Excluded from the generic keyword-based walk below once the object
# inventory has already been extracted from them (or, for subJobInfo/
# object_lineage, once extract_native_report_dependencies() has handled
# them): "job" is one of _CATEGORY_KEYWORDS' substrings (-> "package"), so
# without this exclusion "subJobInfo" would get misclassified as SSIS
# package objects by the generic walker.
_NATIVE_TOP_LEVEL_KEYS = {"inventory", "subJobInfo", "object_lineage"}


def parse_json_report(report_path: Path, result: LakebridgeDiscoveryResult, source_tech: str) -> None:
    with open(report_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    applied = 0
    remaining = data
    if isinstance(data, dict) and isinstance(data.get("inventory"), list):
        applied += _apply_program_inventory_rows(
            result, data["inventory"], source_tech, name_field="name", complexity_field="complexityLevel"
        )
        extract_native_report_dependencies(data, result, source_tech)
        remaining = {k: v for k, v in data.items() if k not in _NATIVE_TOP_LEVEL_KEYS}
    applied += _walk_json_for_lists(remaining, "root", result, source_tech)
    result.raw_report_paths.append(str(report_path))
    if applied == 0:
        result.warnings.append(
            f"lakebridge report {report_path} parsed but no recognizable inventory rows were found "
            f"-- report shape may differ from what report_parser.py expects (unverified mapping)."
        )
    logger.info("Parsed JSON report %s: %d rows classified into inventory categories", report_path, applied)


def parse_excel_report(report_path: Path, result: LakebridgeDiscoveryResult, source_tech: str) -> None:
    try:
        import openpyxl
    except ImportError:
        result.warnings.append(
            f"openpyxl not installed -- cannot parse Excel report {report_path}. "
            f"pip install openpyxl, or rely on --generate-json true instead."
        )
        logger.warning("SKIP excel report %s: openpyxl not installed", report_path)
        return

    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    applied = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        if not header:
            continue
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
        rows = [dict(zip(header, row)) for row in rows_iter]
        if _PROGRAM_NAME_HEADER in header:
            applied += _apply_program_inventory_rows(
                result, rows, source_tech, name_field=_PROGRAM_NAME_HEADER, complexity_field="Complexity"
            )
            continue
        if sheet_name in _NON_INVENTORY_SHEETS:
            continue
        applied += _apply_rows(result, rows, sheet_name, source_tech)
    result.raw_report_paths.append(str(report_path))
    if applied == 0:
        result.warnings.append(
            f"lakebridge report {report_path} parsed but no recognizable inventory rows were found "
            f"-- sheet/column names may differ from what report_parser.py expects (unverified mapping)."
        )
    logger.info("Parsed Excel report %s: %d rows classified into inventory categories", report_path, applied)


def parse_invocation(entity: AnalyzeInvocationEntity, result: LakebridgeDiscoveryResult) -> None:
    """Parses one analyze invocation's report(s) into `result`, in place.
    Never raises: a malformed/missing report is recorded as a warning so it
    doesn't take down the rest of the Lakebridge Discovery run."""
    if entity.status != "success":
        if entity.status != "skipped":
            result.errors.append(f"analyze[{entity.source_tech}] did not succeed: {entity.error}")
        return

    try:
        if entity.report_json_path and Path(entity.report_json_path).exists():
            parse_json_report(Path(entity.report_json_path), result, entity.source_tech)
        elif entity.report_excel_path and Path(entity.report_excel_path).exists():
            parse_excel_report(Path(entity.report_excel_path), result, entity.source_tech)
        else:
            result.warnings.append(f"analyze[{entity.source_tech}] reported success but no report file was found on disk")
    except Exception as exc:  # noqa: BLE001 - a bad report must not crash the whole comparison
        error = f"report_parser failed for analyze[{entity.source_tech}]: {type(exc).__name__}: {exc}"
        result.errors.append(error)
        logger.error(error)
